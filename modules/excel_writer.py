from copy import copy
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


COLUMNS = [
    "Site #",
    "Eval",
    "Site Title",
    "URL",
    "Blurb/Description",
    "Origin (FVEY?)",
    "Malware/Spyware",
    "Tracking",
    "Vulnerability",
    "Content Association",
    "Mainstream",
    "Adversarial",
    "Reputation Check",
    "Final Recommendation",
    "Comments",
    "Date",
]


HEADER_ALIASES = {
    "Site #": ["Site #", "Site Number", "Site No.", "Number"],
    "Eval": ["Eval", "Evaluator", "Initials", "Approver"],
    "Site Title": ["Site Title", "Title", "Website Title"],
    "URL": ["URL", "Website", "Website URL", "Site URL", "Link"],
    "Blurb/Description": ["Blurb/Description", "Description", "Blurb", "Site Description"],
    "Origin (FVEY?)": ["Origin (FVEY?)", "Origin", "FVEY", "Origin/FVEY"],
    "Malware/Spyware": ["Malware/Spyware", "Malware", "Spyware", "Malware Spyware"],
    "Tracking": ["Tracking", "Trackers", "Tracker"],
    "Vulnerability": ["Vulnerability", "Vulnerabilities", "Infrastructure"],
    "Content Association": ["Content Association", "Content", "Association"],
    "Mainstream": ["Mainstream", "Mainstream?"],
    "Adversarial": ["Adversarial", "Adversarial?"],
    "Reputation Check": ["Reputation Check", "Reputation", "Trust Rating"],
    "Final Recommendation": ["Final Recommendation", "Recommendation", "Final"],
    "Comments": ["Comments", "Comment", "Notes"],
    "Date": ["Date", "Date/Time", "Completed", "Date Completed"],
}


def assessment_to_row(assessment: Dict[str, Any], manual: Dict[str, Any]) -> Dict[str, Any]:
    rdap = assessment.get("rdap", {})
    vt = assessment.get("virustotal", {})
    tracking = assessment.get("tracking", {})
    ssl = assessment.get("ssl", {})
    dns = assessment.get("dns", {})
    metadata = assessment.get("metadata", {})
    rec = assessment.get("recommendation", {})

    origin_notes = (
        f"RDAP status: {rdap.get('status', '')}; "
        f"Registrar: {rdap.get('registrar', '')}; "
        f"Country: {rdap.get('country', '')}; "
        f"Source: {rdap.get('source', '')}; "
        f"Notes: {rdap.get('notes', '')}"
    )

    vuln_notes = (
        f"DNS: {dns.get('notes', '')} | "
        f"SSL: {ssl.get('notes', '')}"
    )

    malware_notes = vt.get("notes", "Manual VirusTotal review required.")

    return {
        "Site #": manual.get("site_number", ""),
        "Eval": manual.get("evaluator", ""),
        "Site Title": manual.get("site_title") or metadata.get("title", ""),
        "URL": assessment.get("url", ""),
        "Blurb/Description": manual.get("blurb") or metadata.get("description", ""),
        "Origin (FVEY?)": manual.get("origin") or origin_notes,
        "Malware/Spyware": manual.get("malware") or malware_notes,
        "Tracking": manual.get("tracking") or tracking.get("notes", ""),
        "Vulnerability": manual.get("vulnerability") or vuln_notes,
        "Content Association": manual.get("content_association", "Manual analyst review required."),
        "Mainstream": manual.get("mainstream", "Manual analyst review required."),
        "Adversarial": manual.get("adversarial", "Manual analyst review required."),
        "Reputation Check": manual.get("reputation", "Manual reputation review required."),
        "Final Recommendation": manual.get("final_recommendation") or rec.get("recommendation", "Needs Review"),
        "Comments": manual.get("comments") or rec.get("reason", ""),
        "Date": manual.get("date") or datetime.now().strftime("%Y-%m-%d %H:%M"),
    }


def write_draft_excel(rows: List[Dict[str, Any]], output_path: str) -> str:
    """
    Creates a standalone draft results workbook.
    """
    Path(output_path).parent.mkdir(parents=True, exist_ok=True)

    df = pd.DataFrame(rows, columns=COLUMNS)

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Draft Results")
        ws = writer.book["Draft Results"]

        header_fill = PatternFill("solid", fgColor="1F4E78")
        header_font = Font(color="FFFFFF", bold=True)
        thin = Side(style="thin", color="D9E2F3")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = border

        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
                cell.border = border

        for col in ws.columns:
            max_len = max(len(str(cell.value or "")) for cell in col)
            ws.column_dimensions[col[0].column_letter].width = min(max(max_len + 2, 12), 45)

        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions

    return output_path


def _normalize_header(value: Any) -> str:
    return str(value or "").strip().lower().replace("\n", " ")


def _find_header_row(ws, max_scan_rows: int = 15) -> int:
    """
    Finds the row that appears to contain the checklist headers.
    """
    expected_hits = {
        "url",
        "site title",
        "eval",
        "origin",
        "malware",
        "tracking",
        "vulnerability",
        "content association",
        "mainstream",
        "adversarial",
        "reputation",
        "recommendation",
        "comments",
    }

    best_row = 1
    best_score = 0

    for row_idx in range(1, min(ws.max_row, max_scan_rows) + 1):
        values = [
            _normalize_header(ws.cell(row_idx, col_idx).value)
            for col_idx in range(1, ws.max_column + 1)
        ]

        score = sum(
            1
            for value in values
            if any(hit in value for hit in expected_hits)
        )

        if score > best_score:
            best_score = score
            best_row = row_idx

    return best_row


def _build_header_map(ws, header_row: int) -> Dict[str, int]:
    """
    Maps standard output column names to actual worksheet column numbers.
    """
    sheet_headers: Dict[str, int] = {}

    for col_idx in range(1, ws.max_column + 1):
        raw_value = ws.cell(header_row, col_idx).value
        normalized = _normalize_header(raw_value)
        if normalized:
            sheet_headers[normalized] = col_idx

    mapped: Dict[str, int] = {}

    for standard_name, aliases in HEADER_ALIASES.items():
        for alias in aliases:
            alias_norm = _normalize_header(alias)

            for actual_header, col_idx in sheet_headers.items():
                if alias_norm == actual_header or alias_norm in actual_header:
                    mapped[standard_name] = col_idx
                    break

            if standard_name in mapped:
                break

    return mapped


def _next_blank_row(ws, header_row: int, url_col: Optional[int]) -> int:
    """
    Finds the next row where the URL cell is blank.
    If no URL column exists, appends after the last row.
    """
    if not url_col:
        return ws.max_row + 1

    for row_idx in range(header_row + 1, ws.max_row + 2):
        value = ws.cell(row_idx, url_col).value
        if value is None or str(value).strip() == "":
            return row_idx

    return ws.max_row + 1


def _copy_row_style(ws, source_row: int, target_row: int) -> None:
    """
    Copies basic style from a nearby row into the target row.
    Uses copy() to avoid openpyxl style proxy issues.
    """
    if source_row < 1 or source_row == target_row:
        return

    for col_idx in range(1, ws.max_column + 1):
        source = ws.cell(source_row, col_idx)
        target = ws.cell(target_row, col_idx)

        if source.has_style:
            target.font = copy(source.font)
            target.fill = copy(source.fill)
            target.border = copy(source.border)
            target.number_format = source.number_format
            target.protection = copy(source.protection)
            target.alignment = copy(source.alignment)

        target.alignment = Alignment(
            horizontal=target.alignment.horizontal,
            vertical="top",
            wrap_text=True,
        )

    try:
        ws.row_dimensions[target_row].height = ws.row_dimensions[source_row].height
    except Exception:
        pass


def append_rows_to_template(
    rows: List[Dict[str, Any]],
    template_path: str,
    output_path: str,
    sheet_name: Optional[str] = None,
) -> str:
    """
    Opens the uploaded checklist workbook, appends assessment rows into matching columns,
    and saves a copy. The original template is not modified.
    """
    Path(output_path).parent.mkdir(parents=True, exist_ok=True)

    wb = load_workbook(template_path)

    if sheet_name and sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
    else:
        ws = wb[wb.sheetnames[0]]

    header_row = _find_header_row(ws)
    header_map = _build_header_map(ws, header_row)

    missing_columns = [col for col in COLUMNS if col not in header_map]

    if missing_columns:
        next_col = ws.max_column + 1

        for col_name in missing_columns:
            cell = ws.cell(header_row, next_col)
            cell.value = col_name
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            header_map[col_name] = next_col
            next_col += 1

    url_col = header_map.get("URL")

    for row_data in rows:
        target_row = _next_blank_row(ws, header_row, url_col)

        style_source_row = max(header_row + 1, target_row - 1)
        if style_source_row != target_row:
            _copy_row_style(ws, style_source_row, target_row)

        for standard_col in COLUMNS:
            value = row_data.get(standard_col, "")
            col_idx = header_map.get(standard_col)

            if not col_idx:
                continue

            cell = ws.cell(target_row, col_idx)
            cell.value = value
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    for standard_col, col_idx in header_map.items():
        letter = ws.cell(header_row, col_idx).column_letter

        if standard_col in [
            "Origin (FVEY?)",
            "Malware/Spyware",
            "Tracking",
            "Vulnerability",
            "Content Association",
            "Mainstream",
            "Adversarial",
            "Reputation Check",
            "Comments",
            "Blurb/Description",
        ]:
            ws.column_dimensions[letter].width = max(ws.column_dimensions[letter].width or 10, 32)
        elif standard_col == "URL":
            ws.column_dimensions[letter].width = max(ws.column_dimensions[letter].width or 10, 35)
        else:
            ws.column_dimensions[letter].width = max(ws.column_dimensions[letter].width or 10, 14)

    wb.save(output_path)

    return output_path


__all__ = [
    "COLUMNS",
    "assessment_to_row",
    "write_draft_excel",
    "append_rows_to_template",
]